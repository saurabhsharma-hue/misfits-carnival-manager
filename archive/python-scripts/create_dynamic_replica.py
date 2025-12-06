#!/usr/bin/env python3

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formula.translate import Translator
import re

def read_v2_and_create_dynamic_replica():
    """
    Read OND-JFM Plan with actionables final V2.xlsx and create replica with dynamic sheets
    """

    v2_file = 'OND-JFM Plan with actionbales final V2.xlsx'
    output_file = 'OND-JFM Plan DYNAMIC REPLICA.xlsx'

    print('🔍 Reading V2 file and analyzing structure...')

    try:
        # Read all sheets from V2 file
        xl_file = pd.ExcelFile(v2_file)
        sheet_names = xl_file.sheet_names
        print(f'📊 Found {len(sheet_names)} sheets: {sheet_names}')

        # Read all sheets
        all_sheets = {}
        for sheet_name in sheet_names:
            try:
                all_sheets[sheet_name] = pd.read_excel(v2_file, sheet_name=sheet_name)
                print(f'✅ Read {sheet_name}: {all_sheets[sheet_name].shape}')
            except Exception as e:
                print(f'❌ Error reading {sheet_name}: {e}')

        return all_sheets

    except FileNotFoundError:
        print(f'❌ File not found: {v2_file}')
        return None
    except Exception as e:
        print(f'❌ Error reading file: {e}')
        return None

def parse_club_strategy_for_expansion(working_sheet):
    """
    Parse club strategy column to extract club names for expansion
    """
    print('\n🎯 Parsing Club Strategy column for expansion details...')

    expansion_data = []
    maintenance_data = []

    for idx, row in working_sheet.iterrows():
        activity = row.get('Activity', '')
        city = row.get('City', '')
        area = row.get('Area', '')
        club_strategy = str(row.get('Club_Strategy', '')) if pd.notna(row.get('Club_Strategy')) else ''
        current_clubs = row.get('Current_Clubs_Count', 0) if pd.notna(row.get('Current_Clubs_Count')) else 0
        clubs_needed = row.get('Clubs_Needed_Feb', 0) if pd.notna(row.get('Clubs_Needed_Feb')) else 0

        # Extract club names from strategy text
        club_names = extract_club_names_from_strategy(club_strategy)

        new_clubs_needed = max(0, clubs_needed - current_clubs)

        if new_clubs_needed > 0 and club_names:
            for club_name in club_names:
                expansion_data.append({
                    'Activity': activity,
                    'City': city,
                    'Area': area,
                    'Club_Name': club_name,
                    'Strategy': club_strategy,
                    'Current_Clubs': current_clubs,
                    'Clubs_Needed': clubs_needed,
                    'New_Clubs_Required': new_clubs_needed,
                    'Type': 'Expansion'
                })

        if current_clubs > 0:
            maintenance_data.append({
                'Activity': activity,
                'City': city,
                'Area': area,
                'Club_Names': ', '.join(club_names) if club_names else 'Existing clubs',
                'Current_Clubs': current_clubs,
                'Strategy': club_strategy,
                'Type': 'Maintenance'
            })

    print(f'📈 Found {len(expansion_data)} expansion entries')
    print(f'🔧 Found {len(maintenance_data)} maintenance entries')

    return expansion_data, maintenance_data

def extract_club_names_from_strategy(strategy_text):
    """
    Extract club names from strategy text
    """
    club_names = []

    patterns = [
        r'([A-Z][a-zA-Z\s]+?)\s+(?:club|Club)',
        r'Launch\s+([A-Z][a-zA-Z\s]+?)\s+',
        r'Expand\s+([A-Z][a-zA-Z\s]+?)\s+',
        r'Scale\s+([A-Z][a-zA-Z\s]+?)\s+',
        r'([A-Z][a-zA-Z]+(?:\s+[A-Z][a-zA-Z]+)*)',
    ]

    for pattern in patterns:
        matches = re.findall(pattern, strategy_text)
        for match in matches:
            clean_name = match.strip()
            if len(clean_name) > 2 and clean_name not in ['Launch', 'Expand', 'Scale', 'Club', 'Area', 'Month']:
                club_names.append(clean_name)

    if not club_names and any(keyword in strategy_text.lower() for keyword in ['launch', 'expand', 'scale']):
        club_names = ['Primary Club']

    return list(set(club_names))

def generate_specific_action_from_strategy(original_action, club_name, activity, area):
    """
    Generate specific action based on club strategy and club name
    """
    if club_name == 'TBD' or not club_name:
        return original_action

    if 'expand' in str(original_action).lower():
        if 'meetups per week' in str(original_action):
            action = f"Expand {club_name} ({activity}) in {area} - {original_action}"
        else:
            action = f"Expand {club_name} club in {area} to increase capacity and frequency"
    elif 'increase' in str(original_action).lower():
        action = f"Scale {club_name} ({activity}) in {area} - {original_action}"
    elif 'maintain' in str(original_action).lower():
        action = f"Maintain {club_name} ({activity}) performance in {area} - {original_action}"
    elif 'launch' in str(original_action).lower():
        action = f"Launch new {club_name} chapter in {area} - {original_action}"
    else:
        action = f"Optimize {club_name} ({activity}) in {area} - {original_action}"

    return action

def create_dynamic_replica(all_sheets, expansion_data, maintenance_data):
    """
    Create replica with dynamic Weekly_Execution and Milestones
    """
    print('\n📝 Creating dynamic replica with all sheets...')

    wb = Workbook()
    wb.remove(wb.active)

    # Create all sheets
    for sheet_name, df in all_sheets.items():
        if sheet_name == 'Club_Expansions':
            create_enhanced_club_expansion_sheet(wb, expansion_data, df)
        elif sheet_name == 'Weekly_Execution':
            create_dynamic_weekly_execution_sheet(wb, df)
        elif sheet_name == 'Milestones':
            create_dynamic_milestones_sheet(wb, df)
        else:
            ws = wb.create_sheet(title=sheet_name)
            copy_dataframe_to_sheet(ws, df)

    # Create new Club Maintenance sheet
    create_club_maintenance_sheet(wb, maintenance_data)

    # Save the file
    output_file = 'OND-JFM Plan DYNAMIC REPLICA.xlsx'
    wb.save(output_file)
    print(f'✅ Saved dynamic replica to: {output_file}')

    return output_file

def create_enhanced_club_expansion_sheet(wb, expansion_data, original_club_expansions):
    """
    Create enhanced club expansion sheet with club names and specific actions
    """
    ws = wb.create_sheet(title='Club_Expansions')

    headers = ['Action ID', 'Week Group', 'Type', 'Priority', 'City', 'Area', 'Activity',
              'Specific Action', 'Success Criteria', 'Revenue Impact (₹)', 'Status',
              'Target Date', 'Duration (weeks)', 'Owner', 'Dependencies', 'Strategy Notes',
              'Owner.1', 'Club Name']

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)

    # Copy all original data and enhance
    for row_idx, (_, original_row) in enumerate(original_club_expansions.iterrows(), 2):
        for col_idx, col_name in enumerate(original_club_expansions.columns, 1):
            if col_idx <= 17:
                ws.cell(row=row_idx, column=col_idx, value=original_row[col_name])

        # Add club name and enhanced specific action
        area = original_row.get('Area', '')
        activity = original_row.get('Activity', '')

        club_name = 'TBD'
        for exp_data in expansion_data:
            if (exp_data['Area'] == area and exp_data['Activity'] == activity):
                club_name = exp_data['Club_Name']
                break

        ws.cell(row=row_idx, column=18, value=club_name)

        specific_action = generate_specific_action_from_strategy(
            original_row.get('Specific Action', ''),
            club_name,
            activity,
            area
        )
        ws.cell(row=row_idx, column=8, value=specific_action)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_dynamic_weekly_execution_sheet(wb, original_weekly):
    """
    Create dynamic Weekly_Execution sheet with formulas
    """
    ws = wb.create_sheet(title='Weekly_Execution')

    # Headers
    headers = ['Week', 'Dates', 'Month', 'Expansion Total', 'Expansion Done', 'Expansion %',
              'Launch Total', 'Launch Done', 'Launch %', 'Overall Total', 'Overall Done',
              'Overall %', 'Weekly Revenue Impact (₹)', 'Key Activities This Week']

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)

    # Copy basic data and add dynamic formulas
    for row_idx, (_, original_row) in enumerate(original_weekly.iterrows(), 2):
        # Basic data columns
        ws.cell(row=row_idx, column=1, value=original_row['Week'])
        ws.cell(row=row_idx, column=2, value=original_row['Dates'])
        ws.cell(row=row_idx, column=3, value=original_row['Month'])

        # Dynamic formulas for tracking
        week_group = f"'{original_row['Month']} 2024 (Weeks {row_idx-1}-{min(row_idx+2, 18)})'"

        # Expansion metrics with COUNTIFS
        ws.cell(row=row_idx, column=4, value=f'=COUNTIFS(Club_Expansions[Week Group],"{week_group}")')
        ws.cell(row=row_idx, column=5, value=f'=COUNTIFS(Club_Expansions[Week Group],"{week_group}",Club_Expansions[Status],"COMPLETED")')
        ws.cell(row=row_idx, column=6, value=f'=IF(D{row_idx}=0,0,E{row_idx}/D{row_idx})')

        # Launch metrics with COUNTIFS
        ws.cell(row=row_idx, column=7, value=f'=COUNTIFS(Club_Launches[Week Group],"{week_group}")')
        ws.cell(row=row_idx, column=8, value=f'=COUNTIFS(Club_Launches[Week Group],"{week_group}",Club_Launches[Status],"COMPLETED")')
        ws.cell(row=row_idx, column=9, value=f'=IF(G{row_idx}=0,0,H{row_idx}/G{row_idx})')

        # Overall metrics
        ws.cell(row=row_idx, column=10, value=f'=D{row_idx}+G{row_idx}')
        ws.cell(row=row_idx, column=11, value=f'=E{row_idx}+H{row_idx}')
        ws.cell(row=row_idx, column=12, value=f'=IF(J{row_idx}=0,0,K{row_idx}/J{row_idx})')

        # Revenue impact
        ws.cell(row=row_idx, column=13, value=f'=SUMIFS(Club_Expansions[Revenue Impact (₹)],Club_Expansions[Week Group],"{week_group}",Club_Expansions[Status],"COMPLETED")+SUMIFS(Club_Launches[Revenue Impact (₹)],Club_Launches[Week Group],"{week_group}",Club_Launches[Status],"COMPLETED")')

        # Key activities
        ws.cell(row=row_idx, column=14, value=f'Week {row_idx-1} focus areas')

    # Format percentage columns
    for row in range(2, 20):
        for col in [6, 9, 12]:  # Percentage columns
            cell = ws.cell(row=row, column=col)
            cell.number_format = '0.0%'

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_dynamic_milestones_sheet(wb, original_milestones):
    """
    Create dynamic Milestones sheet with formulas
    """
    ws = wb.create_sheet(title='Milestones')

    # Headers
    headers = ['Monthly Milestone', 'Target Date', 'Week', 'Description', 'Revenue Target (₹)',
              'Status', 'Key Cities', 'Actions Focus', 'Club Target', 'Attendance Target',
              'Quality Metric', 'Team Focus', 'Total Actions', 'Completed', 'Progress %', 'Notes']

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='9BBB59', end_color='9BBB59', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)

    # Copy data and add dynamic formulas
    for row_idx, (_, original_row) in enumerate(original_milestones.iterrows(), 2):
        # Basic data columns (1-12)
        for col_idx in range(1, 13):
            col_name = original_milestones.columns[col_idx-1]
            ws.cell(row=row_idx, column=col_idx, value=original_row[col_name])

        # Dynamic formulas for tracking
        week_num = original_row.get('Week', row_idx-1)

        # Total Actions (sum of expansions and launches for this week)
        ws.cell(row=row_idx, column=13, value=f'=COUNTIFS(Club_Expansions[Week],{week_num})+COUNTIFS(Club_Launches[Week],{week_num})')

        # Completed Actions
        ws.cell(row=row_idx, column=14, value=f'=COUNTIFS(Club_Expansions[Week],{week_num},Club_Expansions[Status],"COMPLETED")+COUNTIFS(Club_Launches[Week],{week_num},Club_Launches[Status],"COMPLETED")')

        # Progress %
        ws.cell(row=row_idx, column=15, value=f'=IF(M{row_idx}=0,0,N{row_idx}/M{row_idx})')

        # Notes - dynamic based on progress
        ws.cell(row=row_idx, column=16, value=f'=IF(O{row_idx}>=0.8,"On Track",IF(O{row_idx}>=0.5,"Delayed","Critical"))')

    # Format percentage column
    for row in range(2, 10):
        cell = ws.cell(row=row, column=15)
        cell.number_format = '0.0%'

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_club_maintenance_sheet(wb, maintenance_data):
    """
    Create club maintenance sheet
    """
    ws = wb.create_sheet(title='Club to be Maintained')

    headers = ['Activity', 'City', 'Area', 'Club Names', 'Current Clubs',
               'Maintenance Strategy', 'Focus Area', 'Expected Outcome']

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='2F5233', end_color='2F5233', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)

    # Add data
    for row_idx, data in enumerate(maintenance_data, 2):
        ws.cell(row=row_idx, column=1, value=data['Activity'])
        ws.cell(row=row_idx, column=2, value=data['City'])
        ws.cell(row=row_idx, column=3, value=data['Area'])
        ws.cell(row=row_idx, column=4, value=data['Club_Names'])
        ws.cell(row=row_idx, column=5, value=data['Current_Clubs'])

        if data['Activity'] in ['MUSIC', 'BOARDGAMING', 'SOCIAL_DEDUCTIONS']:
            strategy = 'Maximize engagement and frequency'
        else:
            strategy = 'Maintain quality and consistency'
        ws.cell(row=row_idx, column=6, value=strategy)

        ws.cell(row=row_idx, column=7, value='Member retention and activity scaling')
        ws.cell(row=row_idx, column=8, value=f'Sustain {data["Current_Clubs"]} active clubs')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

def copy_dataframe_to_sheet(ws, df):
    """
    Copy dataframe to worksheet
    """
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Format headers
    if ws.max_row > 0:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

def main():
    print('🚀 CREATING DYNAMIC REPLICA WITH FORMULAS')
    print('=' * 60)

    # Read V2 file
    all_sheets = read_v2_and_create_dynamic_replica()

    if all_sheets is None:
        print('❌ Could not read sheets')
        return

    # Parse club strategy
    working_sheet = all_sheets['Working sheet']
    expansion_data, maintenance_data = parse_club_strategy_for_expansion(working_sheet)

    # Create dynamic replica
    output_file = create_dynamic_replica(all_sheets, expansion_data, maintenance_data)

    print(f'\n✅ DYNAMIC REPLICA CREATED: {output_file}')
    print('📊 Features added:')
    print('   • Enhanced Club_Expansions with club names in column H')
    print('   • Dynamic Weekly_Execution with COUNTIFS formulas')
    print('   • Dynamic Milestones with progress tracking')
    print('   • Club Maintenance sheet')
    print('   • All formulas reference Club_Expansions and Club_Launches sheets')

if __name__ == "__main__":
    main()