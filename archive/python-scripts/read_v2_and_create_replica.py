#!/usr/bin/env python3

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re

def read_v2_and_create_replica():
    """
    Read OND-JFM Plan with actionables final V2.xlsx and create replica with all sheets
    """

    v2_file = 'OND-JFM Plan with actionbales final V2.xlsx'
    output_file = 'OND-JFM Plan REPLICA with Club Maintenance.xlsx'

    print('🔍 Reading V2 file and analyzing structure...')

    try:
        # Read all sheets from V2 file
        xl_file = pd.ExcelFile(v2_file)
        sheet_names = xl_file.sheet_names
        print(f'📊 Found {len(sheet_names)} sheets: {sheet_names}')

        # Read working sheet first
        working_sheet = pd.read_excel(v2_file, sheet_name='Working sheet')
        print(f'✅ Working sheet loaded: {working_sheet.shape[0]} rows, {working_sheet.shape[1]} columns')

        # Display column names
        print(f'📋 Columns: {list(working_sheet.columns)}')

        # Read all other sheets
        all_sheets = {}
        for sheet_name in sheet_names:
            try:
                all_sheets[sheet_name] = pd.read_excel(v2_file, sheet_name=sheet_name)
                print(f'✅ Read {sheet_name}: {all_sheets[sheet_name].shape}')
            except Exception as e:
                print(f'❌ Error reading {sheet_name}: {e}')

        return all_sheets, working_sheet

    except FileNotFoundError:
        print(f'❌ File not found: {v2_file}')
        return None, None
    except Exception as e:
        print(f'❌ Error reading file: {e}')
        return None, None

def parse_club_strategy_for_expansion(working_sheet):
    """
    Parse club strategy column to extract club names for expansion
    """
    print('\n🎯 Parsing Club Strategy column for expansion details...')

    expansion_data = []
    maintenance_data = []

    for idx, row in working_sheet.iterrows():
        activity = row.get('Activity', '')
        city = row.get('City', '')
        area = row.get('Area', '')
        club_strategy = str(row.get('Club_Strategy', '')) if pd.notna(row.get('Club_Strategy')) else ''
        current_clubs = row.get('Current_Clubs_Count', 0) if pd.notna(row.get('Current_Clubs_Count')) else 0
        clubs_needed = row.get('Clubs_Needed_Feb', 0) if pd.notna(row.get('Clubs_Needed_Feb')) else 0

        # Extract club names from strategy text
        club_names = extract_club_names_from_strategy(club_strategy)

        new_clubs_needed = max(0, clubs_needed - current_clubs)

        if new_clubs_needed > 0 and club_names:
            # This needs expansion
            for club_name in club_names:
                expansion_data.append({
                    'Activity': activity,
                    'City': city,
                    'Area': area,
                    'Club_Name': club_name,
                    'Strategy': club_strategy,
                    'Current_Clubs': current_clubs,
                    'Clubs_Needed': clubs_needed,
                    'New_Clubs_Required': new_clubs_needed,
                    'Type': 'Expansion'
                })

        if current_clubs > 0:
            # These need maintenance
            maintenance_data.append({
                'Activity': activity,
                'City': city,
                'Area': area,
                'Club_Names': ', '.join(club_names) if club_names else 'Existing clubs',
                'Current_Clubs': current_clubs,
                'Strategy': club_strategy,
                'Type': 'Maintenance'
            })

    print(f'📈 Found {len(expansion_data)} expansion entries')
    print(f'🔧 Found {len(maintenance_data)} maintenance entries')

    return expansion_data, maintenance_data

def extract_club_names_from_strategy(strategy_text):
    """
    Extract club names from strategy text
    """
    club_names = []

    # Common patterns to look for club names
    patterns = [
        r'([A-Z][a-zA-Z\s]+?)\s+(?:club|Club)',  # "GameMasters club"
        r'Launch\s+([A-Z][a-zA-Z\s]+?)\s+',       # "Launch Ballers "
        r'Expand\s+([A-Z][a-zA-Z\s]+?)\s+',       # "Expand Ballers "
        r'Scale\s+([A-Z][a-zA-Z\s]+?)\s+',        # "Scale GameMasters "
        r'([A-Z][a-zA-Z]+(?:\s+[A-Z][a-zA-Z]+)*)',  # General capitalized words
    ]

    for pattern in patterns:
        matches = re.findall(pattern, strategy_text)
        for match in matches:
            clean_name = match.strip()
            if len(clean_name) > 2 and clean_name not in ['Launch', 'Expand', 'Scale', 'Club', 'Area', 'Month']:
                club_names.append(clean_name)

    # If no specific names found, create generic ones based on activity
    if not club_names and any(keyword in strategy_text.lower() for keyword in ['launch', 'expand', 'scale']):
        club_names = ['Primary Club']

    return list(set(club_names))  # Remove duplicates

def generate_specific_action_from_strategy(original_action, club_name, activity, area):
    """
    Generate specific action based on club strategy and club name
    """
    if club_name == 'TBD' or not club_name:
        return original_action

    # Create enhanced specific action with club name
    if 'expand' in original_action.lower():
        if 'meetups per week' in original_action:
            action = f"Expand {club_name} ({activity}) in {area} - {original_action}"
        else:
            action = f"Expand {club_name} club in {area} to increase capacity and frequency"
    elif 'increase' in original_action.lower():
        action = f"Scale {club_name} ({activity}) in {area} - {original_action}"
    elif 'maintain' in original_action.lower():
        action = f"Maintain {club_name} ({activity}) performance in {area} - {original_action}"
    elif 'launch' in original_action.lower():
        action = f"Launch new {club_name} chapter in {area} - {original_action}"
    else:
        # Default enhancement
        action = f"Optimize {club_name} ({activity}) in {area} - {original_action}"

    return action

def create_replica_with_maintenance(all_sheets, working_sheet, expansion_data, maintenance_data):
    """
    Create replica of all sheets with club maintenance sheet
    """
    print('\n📝 Creating replica with all sheets...')

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create all original sheets first
    for sheet_name, df in all_sheets.items():
        if sheet_name == 'Club_Expansions' and expansion_data:
            # Enhanced club expansion with club names
            create_enhanced_club_expansion_sheet(wb, expansion_data, df)
        else:
            # Copy original sheet
            ws = wb.create_sheet(title=sheet_name)
            copy_dataframe_to_sheet(ws, df)

    # Create new Club Maintenance sheet
    create_club_maintenance_sheet(wb, maintenance_data)

    # Save the file
    output_file = 'OND-JFM Plan REPLICA with Club Maintenance.xlsx'
    wb.save(output_file)
    print(f'✅ Saved replica to: {output_file}')

    return output_file

def create_enhanced_club_expansion_sheet(wb, expansion_data, original_club_expansions):
    """
    Create enhanced club expansion sheet with club names and specific actions
    """
    ws = wb.create_sheet(title='Club_Expansions')

    # Use original headers from V2 file
    headers = ['Action ID', 'Week Group', 'Type', 'Priority', 'City', 'Area', 'Activity',
              'Specific Action', 'Success Criteria', 'Revenue Impact (₹)', 'Status',
              'Target Date', 'Duration (weeks)', 'Owner', 'Dependencies', 'Strategy Notes',
              'Owner.1', 'Club Name']

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)

    # First copy all original data
    for row_idx, (_, original_row) in enumerate(original_club_expansions.iterrows(), 2):
        for col_idx, col_name in enumerate(original_club_expansions.columns, 1):
            if col_idx <= 17:  # Original columns
                ws.cell(row=row_idx, column=col_idx, value=original_row[col_name])

        # Add club name in column 18 based on strategy
        area = original_row.get('Area', '')
        activity = original_row.get('Activity', '')

        # Find matching club name from expansion data
        club_name = 'TBD'
        for exp_data in expansion_data:
            if (exp_data['Area'] == area and exp_data['Activity'] == activity):
                club_name = exp_data['Club_Name']
                break

        ws.cell(row=row_idx, column=18, value=club_name)

        # Update Specific Action (Column H) with detailed strategy
        specific_action = generate_specific_action_from_strategy(
            original_row.get('Specific Action', ''),
            club_name,
            activity,
            area
        )
        ws.cell(row=row_idx, column=8, value=specific_action)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_club_maintenance_sheet(wb, maintenance_data):
    """
    Create club maintenance sheet
    """
    ws = wb.create_sheet(title='Club to be Maintained')

    # Headers
    headers = ['Activity', 'City', 'Area', 'Club Names', 'Current Clubs',
               'Maintenance Strategy', 'Focus Area', 'Expected Outcome']

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='2F5233', end_color='2F5233', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)

    # Add data
    for row_idx, data in enumerate(maintenance_data, 2):
        ws.cell(row=row_idx, column=1, value=data['Activity'])
        ws.cell(row=row_idx, column=2, value=data['City'])
        ws.cell(row=row_idx, column=3, value=data['Area'])
        ws.cell(row=row_idx, column=4, value=data['Club_Names'])
        ws.cell(row=row_idx, column=5, value=data['Current_Clubs'])

        # Generate maintenance strategy
        if data['Activity'] in ['MUSIC', 'BOARDGAMING', 'SOCIAL_DEDUCTIONS']:
            strategy = 'Maximize engagement and frequency'
        else:
            strategy = 'Maintain quality and consistency'
        ws.cell(row=row_idx, column=6, value=strategy)

        # Focus area
        focus = 'Member retention and activity scaling'
        ws.cell(row=row_idx, column=7, value=focus)

        # Expected outcome
        outcome = f'Sustain {data["Current_Clubs"]} active clubs'
        ws.cell(row=row_idx, column=8, value=outcome)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def copy_dataframe_to_sheet(ws, df):
    """
    Copy dataframe to worksheet
    """
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Format headers
    if ws.max_row > 0:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

def main():
    print('🚀 CREATING REPLICA WITH CLUB MAINTENANCE SHEET')
    print('=' * 60)

    # Read V2 file
    all_sheets, working_sheet = read_v2_and_create_replica()

    if working_sheet is None:
        print('❌ Could not read working sheet')
        return

    # Parse club strategy
    expansion_data, maintenance_data = parse_club_strategy_for_expansion(working_sheet)

    # Create replica
    output_file = create_replica_with_maintenance(all_sheets, working_sheet, expansion_data, maintenance_data)

    print(f'\n✅ REPLICA CREATED SUCCESSFULLY: {output_file}')
    print(f'📊 Club Expansion entries: {len(expansion_data)}')
    print(f'🔧 Club Maintenance entries: {len(maintenance_data)}')

    # Display sample data
    if expansion_data:
        print('\n🎯 SAMPLE CLUB EXPANSION DATA:')
        for i, data in enumerate(expansion_data[:5], 1):
            print(f'{i}. {data["Activity"]} - {data["Club_Name"]} in {data["Area"]}, {data["City"]}')

    if maintenance_data:
        print('\n🔧 SAMPLE CLUB MAINTENANCE DATA:')
        for i, data in enumerate(maintenance_data[:5], 1):
            print(f'{i}. {data["Activity"]} - {data["Club_Names"]} in {data["Area"]}, {data["City"]}')

if __name__ == "__main__":
    main()